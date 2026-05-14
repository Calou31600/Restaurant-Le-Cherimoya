import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def create_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Evenements"

    # En-têtes
    headers = [
        "Nom", "Theme", "Date_Debut", "Date_Fin", 
        "Actif", "Prix", "Sous_Titre", 
        "Entrees", "Plats", "Desserts", "Photo"
    ]
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid") # Violet pour les événements
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        ws.column_dimensions[cell.column_letter].width = 20

    # Largeur spécifique pour les listes
    ws.column_dimensions['H'].width = 40 # Entrees
    ws.column_dimensions['I'].width = 40 # Plats
    ws.column_dimensions['J'].width = 40 # Desserts

    # Exemple de données
    ws.append([
        "Soirée Saint-Valentin", "fete", "2026-02-14", "2026-02-14", 
        "VRAI", "45€", "Un moment romantique inoubliable", 
        "Coupe de champagne\nFoie gras maison", 
        "Filet mignon aux morilles\nRisotto aux asperges", 
        "Cœur fondant au chocolat\nCafé gourmand", 
        "" # Photo
    ])
    
    # Ajuster la hauteur des lignes
    for i in range(2, 10):
        ws.row_dimensions[i].height = 80

    filename = "modele_import_evenements.xlsx"
    wb.save(filename)
    print(f"✅ Modèle Excel créé : {filename}")

if __name__ == "__main__":
    create_template()
