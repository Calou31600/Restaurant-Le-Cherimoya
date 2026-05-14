import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def create_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Vins"

    # En-têtes
    headers = [
        "Nom", "Nom_EN", "Appellation", "Millesime", 
        "Prix_Verre", "Prix_Bouteille", "Type", 
        "Description", "Description_EN", "Photo (Insérer l'image dans cette cellule)"
    ]
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        ws.column_dimensions[cell.column_letter].width = 25

    # Exemple de données
    ws.append([
        "Château Margaux", "Margaux Castle", "Margaux", "2015", 
        "25.00", "150.00", "Rouge", 
        "Un vin élégant et complexe.", "An elegant and complex wine.", 
        "" # Laisser vide pour l'image
    ])
    
    # Ajuster la hauteur des lignes pour les images
    for i in range(2, 10):
        ws.row_dimensions[i].height = 60

    filename = "modele_import_vins.xlsx"
    wb.save(filename)
    print(f"✅ Modèle Excel créé : {filename}")

if __name__ == "__main__":
    create_template()
