import os
import sys
import json
import requests
import cloudinary
import cloudinary.uploader
from io import BytesIO
from dotenv import load_dotenv
from openpyxl import load_workbook

# Ajouter le dossier parent au path pour les imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

load_dotenv()

class EventImporter:
    def __init__(self):
        self.api_key = os.getenv('AIRTABLE_API_KEY')
        self.base_id = os.getenv('AIRTABLE_BASE_ID')
        self.headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        self.setup_cloudinary()

    def setup_cloudinary(self):
        cloudinary_url = os.getenv("CLOUDINARY_URL", "")
        if cloudinary_url.startswith("cloudinary://"):
            try:
                creds = cloudinary_url.replace("cloudinary://", "")
                key_secret, cloud_name = creds.split("@")
                api_key, api_secret = key_secret.split(":")
                cloudinary.config(cloud_name=cloud_name, api_key=api_key, api_secret=api_secret)
                print("✅ Cloudinary configuré.")
            except Exception as e:
                print(f"⚠️ Erreur config Cloudinary: {e}")

    def upload_image(self, file_source):
        if not file_source: return ""
        try:
            print(f"📤 Uploading vers Cloudinary...")
            res = cloudinary.uploader.upload(file_source, folder="restaurant_cherimoya/evenements")
            return res.get('secure_url')
        except Exception as e:
            print(f"❌ Erreur upload Cloudinary: {e}")
            return ""

    def import_xlsx(self, xlsx_path):
        if not os.path.exists(xlsx_path):
            print(f"❌ Fichier non trouvé: {xlsx_path}")
            return

        print(f"🚀 Début de l'importation depuis {xlsx_path}...")
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        images_by_row = {}
        for image in ws._images:
            row_idx = image.anchor._from.row + 1
            images_by_row[row_idx] = image.ref

        headers = [cell.value for cell in ws[1]]
        records = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                if header:
                    row_data[header] = ws.cell(row=row_idx, column=col_idx).value
            
            if not row_data.get('Nom'): continue
            
            print(f"🔍 Traitement de: {row_data.get('Nom')}")
            
            photo_url = ""
            if row_idx in images_by_row:
                img_data = images_by_row[row_idx]
                photo_url = self.upload_image(BytesIO(img_data.read()))
            
            # Formatage pour Airtable
            fields = {
                "Name":       row_data.get('Nom', ''),
                "Theme":      row_data.get('Theme', 'personnalise'),
                "Start_Date": str(row_data.get('Date_Debut', '')),
                "End_Date":   str(row_data.get('Date_Fin', '')),
                "Active":     str(row_data.get('Actif', '')).upper() in ['VRAI', 'TRUE', '1', 'YES'],
                "Price":      str(row_data.get('Prix', '')),
                "Subtitle":   str(row_data.get('Sous_Titre', '')),
                "Entrees":    str(row_data.get('Entrees', '')),
                "Plats":      str(row_data.get('Plats', '')),
                "Desserts":   str(row_data.get('Desserts', '')),
                "Photo":      photo_url
            }
            
            records.append({"fields": fields})
            
            if len(records) >= 10:
                self.push_to_airtable(records)
                records = []
        
        if records:
            self.push_to_airtable(records)

    def push_to_airtable(self, records):
        url = f"https://api.airtable.com/v0/{self.base_id}/Special_Menus_Cherimoya"
        try:
            response = requests.post(url, headers=self.headers, json={"records": records, "typecast": True})
            if response.status_code == 200:
                print(f"✅ {len(records)} événements ajoutés.")
            else:
                print(f"❌ Erreur Airtable [{response.status_code}]: {response.text}")
        except Exception as e:
            print(f"❌ Exception Airtable: {e}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_evenements_excel.py <fichier.xlsx>")
    else:
        importer = EventImporter()
        importer.import_xlsx(sys.argv[1])
