import os
import csv
import sys
import json
import requests
import cloudinary
import cloudinary.uploader
from io import BytesIO
from dotenv import load_dotenv

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Ajouter le dossier parent au path pour les imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

load_dotenv()

class WineImporter:
    def __init__(self):
        self.api_key = os.getenv('AIRTABLE_API_KEY')
        self.base_id = os.getenv('AIRTABLE_BASE_ID')
        self.headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        self.setup_cloudinary()

    def setup_cloudinary(self):
        cloudinary_url = os.getenv("CLOUDINARY_URL", "")
        if cloudinary_url.startswith("cloudinary://"):
            try:
                creds = cloudinary_url.replace("cloudinary://", "")
                key_secret, cloud_name = creds.split("@")
                api_key, api_secret = key_secret.split(":")
                cloudinary.config(cloud_name=cloud_name, api_key=api_key, api_secret=api_secret)
                print("✅ Cloudinary configuré.")
            except Exception as e:
                print(f"⚠️ Erreur config Cloudinary: {e}")

    def upload_image(self, file_source, filename="wine_image.jpg"):
        """Upload une image vers Cloudinary (chemin local, URL ou Bytes)."""
        if not file_source:
            return ""
        
        if isinstance(file_source, str) and file_source.startswith('http'):
            return file_source
        
        try:
            print(f"📤 Uploading vers Cloudinary...")
            res = cloudinary.uploader.upload(file_source, folder="restaurant_cherimoya/vins")
            return res.get('secure_url')
        except Exception as e:
            print(f"❌ Erreur upload Cloudinary: {e}")
            return ""

    def import_any(self, file_path):
        if not os.path.exists(file_path):
            print(f"❌ Fichier non trouvé: {file_path}")
            return

        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            self.import_csv(file_path)
        elif ext == '.xlsx':
            if not HAS_OPENPYXL:
                print("❌ openpyxl n'est pas installé. Utilisez 'pip install openpyxl'.")
                return
            self.import_xlsx(file_path)
        else:
            print(f"❌ Extension non supportée: {ext}")

    def import_csv(self, csv_path):
        print(f"🚀 Début de l'importation depuis CSV {csv_path}...")
        with open(csv_path, mode='r', encoding='utf-8') as f:
            content = f.read(2048)
            f.seek(0)
            dialect = csv.Sniffer().sniff(content) if content else csv.excel
            reader = csv.DictReader(f, dialect=dialect)
            
            records = []
            for row in reader:
                photo_url = self.upload_image(row.get('Photo_URL_ou_Chemin', ''))
                records.append({"fields": self.build_fields(row, photo_url)})
                if len(records) >= 10:
                    self.push_to_airtable(records)
                    records = []
            if records: self.push_to_airtable(records)

    def import_xlsx(self, xlsx_path):
        print(f"🚀 Début de l'importation depuis Excel {xlsx_path}...")
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        # 1. Mapper les images par ligne
        # Attention: image.anchor._from.row est 0-indexed, donc l'image sur la ligne 2 a row=1
        images_by_row = {}
        for image in ws._images:
            row_idx = image.anchor._from.row + 1 # Convertir en 1-indexed pour correspondre à ws.cell
            images_by_row[row_idx] = image.ref # Données binaires de l'image

        # 2. Lire les données
        headers = [cell.value for cell in ws[1]]
        records = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                if header:
                    row_data[header] = ws.cell(row=row_idx, column=col_idx).value
            
            if not row_data.get('Nom'): continue
            
            print(f"🔍 Traitement de: {row_data.get('Nom')}")
            
            # Récupérer l'image si elle existe pour cette ligne
            photo_url = ""
            if row_idx in images_by_row:
                img_data = images_by_row[row_idx]
                photo_url = self.upload_image(BytesIO(img_data.read()))
            
            records.append({"fields": self.build_fields(row_data, photo_url)})
            
            if len(records) >= 10:
                self.push_to_airtable(records)
                records = []
        
        if records: self.push_to_airtable(records)

    def build_fields(self, data, photo_url):
        def format_price(val):
            if not val: return ""
            try:
                if isinstance(val, str):
                    val = val.replace(',', '.').strip()
                return f"€{float(val):.2f}"
            except:
                return str(val)

        return {
            "Nom": data.get('Nom', '') or '',
            "Nom_EN": data.get('Nom_EN', '') or '',
            "Appellation": data.get('Appellation', '') or '',
            "Millesime": str(data.get('Millesime', '') or ''),
            "Prix_Verre": format_price(data.get('Prix_Verre', '')),
            "Prix_Bouteille": format_price(data.get('Prix_Bouteille', '')),
            "Type": data.get('Type', '') or '',
            "Description": data.get('Description', '') or '',
            "Description_EN": data.get('Description_EN', '') or '',
            "Photo": photo_url
        }

    def push_to_airtable(self, records):
        url = f"https://api.airtable.com/v0/{self.base_id}/Carte_Vins"
        try:
            response = requests.post(url, headers=self.headers, json={"records": records, "typecast": True})
            if response.status_code == 200:
                print(f"✅ {len(records)} vins ajoutés.")
            else:
                print(f"❌ Erreur Airtable [{response.status_code}]: {response.text}")
        except Exception as e:
            print(f"❌ Exception Airtable: {e}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_vins_excel.py <fichier.csv ou fichier.xlsx>")
    else:
        importer = WineImporter()
        importer.import_any(sys.argv[1])
